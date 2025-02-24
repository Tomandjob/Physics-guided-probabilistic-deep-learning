import numpy as np
import tensorflow as tf
import tensorflow_probability as tfp
from scipy.optimize import minimize
from openpyxl import Workbook  
tfd = tfp.distributions
tfpl = tfp.layers
TF_ENABLE_ONEDNN_OPTS=0
# Define the neural network with Bayesian layers
class BayesianPINN(tf.keras.Model):
    def __init__(self):
        super(BayesianPINN, self).__init__()
        #self.dense1 = tfpl.DenseReparameterization(5, activation=tf.nn.tanh)
        self.dense1 = tf.keras.layers.Dense(50, activation=tf.nn.tanh)
        self.dense2 = tf.keras.layers.Dense(50, activation=tf.nn.tanh)
 #       self.dense3 = tf.keras.layers.Dense(50, activation=tf.nn.tanh)
        self.dense4 = tfpl.DenseReparameterization(1)

    def call(self, inputs):
        x = inputs
        x = self.dense1(x)
        x = self.dense2(x)
  #      x = self.dense3(x)
        x = self.dense4(x)
        return x

# Define the loss function
def loss(model1, x, y, boundary_x, boundary_y):
    with tf.GradientTape() as tape:
        tape.watch(x)  # Ensure that x is watched by the tape for gradient calculation
        y_pred = model1(x)
        f_pred = tape.gradient(y_pred, x)  # Derivative of y_pred w.r.t. x
        # print("x:", x.shape)
        # print("y_pred_dist:", y_pred.shape)
    # Define the differential equation
    f = f_pred - (0.018849624019951 * y_pred + 3.473893718622038) / x
    mse = tf.reduce_mean(tf.square(f))  # MSE loss for the differential equation
    boundary_loss = tf.reduce_mean(tf.square(model1(boundary_x) - boundary_y))  # Boundary condition loss

    return mse + boundary_loss


# Define the boundary condition
boundary_x = tf.constant([[45]], dtype=tf.float32)
boundary_y = tf.constant([[0]], dtype=tf.float32)

# Generate training data
x_train = np.linspace(10, 60, 40).reshape(-1, 1).astype(np.float32)
x_train = tf.constant(x_train)
y_train = np.zeros_like(x_train)

# Initialize BayesianPINN
model1 = BayesianPINN()

# Define optimization step
optimizer = tf.keras.optimizers.Adam()

# Training loop
epochs = 20000
for epoch in range(epochs):
    with tf.GradientTape() as tape:
        loss_value = loss(model1, x_train, y_train, boundary_x, boundary_y)
    gradients = tape.gradient(loss_value, model1.trainable_variables)
    optimizer.apply_gradients(zip(gradients, model1.trainable_variables))
    if epoch % 100 == 0:
        print("Epoch {}: Loss {}".format(epoch, loss_value.numpy()))
#
# # 保存模型权重
model1.save_weights('model1.weights.h5')


# 至少调用一次模型，以便 TensorFlow 创建模型的变量
_ = model1(x_train[:1])
# 加载模型权重
model1.load_weights('model1.weights.h5')

# Evaluate the trained model
x_test = np.linspace(10, 60, 40).reshape(-1, 1).astype(np.float32)



# 创建Excel工作簿和工作表  
wb = Workbook()  
ws = wb.active  
ws.title = 'sheet1'  
  
# 写入数据到工作表  
for i in range(len(x_test)):  
    # 提取x_test中第i行的第0个元素（即单个值）  
    value = x_test[i, 0]  
    ws.cell(row=i+1, column=1, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）  
  
# 保存Excel文件  
wb.save('W1=0.5.xlsx')


x_test= tf.constant(x_test)
y_test = np.zeros_like(x_test)


for i in range(len(y_test)):  
    # 提取x_test中第i行的第0个元素（即单个值）  
    value = y_test[i, 0]  
    ws.cell(row=i+1, column=2, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）  
  
# 保存Excel文件  
wb.save('W1=0.5.xlsx')


import matplotlib.pyplot as plt
# 运行模型多次以获取每个点的预测样本
num_samples = 1000
predictions = [model1(x_test).numpy() for _ in range(num_samples)]

# 计算每个点的平均值和标准差
mean_predictions = np.mean(predictions, axis=0)
std_predictions = np.std(predictions, axis=0)
print(x_test.shape)
print(mean_predictions.shape)
print(std_predictions.shape)

x_test=x_test.numpy().flatten()
mean_predictions=mean_predictions.flatten()
std_predictions=std_predictions.flatten()

print(mean_predictions)
print(std_predictions)
for i in range(len(mean_predictions)):  
    # 提取x_test中第i行的第0个元素（即单个值）  
    value = mean_predictions[i]  
    ws.cell(row=i+1, column=3, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）  
  
# 保存Excel文件  
wb.save('W1=0.5.xlsx')

for i in range(len(std_predictions)):  
    # 提取x_test中第i行的第0个元素（即单个值）  
    value = std_predictions[i]  
    ws.cell(row=i+1, column=4, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）  
  
# 保存Excel文件  
wb.save('W1=0.5.xlsx')

# 绘制结果
plt.plot(x_test, mean_predictions, label='Predicted Mean')
plt.fill_between(x_test, mean_predictions - 2 * std_predictions, mean_predictions + 2 * std_predictions, alpha=0.3, label='Uncertainty')
plt.scatter(boundary_x, boundary_y, color='red', label='Boundary Condition')
plt.xlabel('x')
plt.ylabel('y')
plt.legend()
plt.show()



#
#
# y_pred_samples = model(x_test).numpy()
#
# # Plot the results
# import matplotlib.pyplot as plt
# plt.plot(x_test, y_pred_samples, label='Predicted')
# plt.scatter(boundary_x, boundary_y, color='red', label='Boundary Condition')
# plt.xlabel('x')
# plt.ylabel('y')
# plt.legend()
# plt.show()





# # Compute mean prediction
# y_pred_mean = tf.reduce_mean(y_pred_samples, axis=0).numpy()

# # Plot the results
# import matplotlib.pyplot as plt
# plt.plot(x_test, np.tile(y_pred_mean, (x_test.shape[0], 1)), label='Predicted')
# plt.scatter(boundary_x, boundary_y, color='red', label='Boundary Condition')
# plt.xlabel('x')
# plt.ylabel('y')
# plt.legend()
# plt.show()
