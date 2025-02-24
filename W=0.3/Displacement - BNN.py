import numpy as np
import tensorflow as tf
import tensorflow_probability as tfp
from scipy.optimize import minimize
import xlwt
from openpyxl import Workbook
tfd = tfp.distributions
tfpl = tfp.layers

R = 10
n0 = 2/1.1
E_c = 31500
q = 5.13
delta_c = 0.8
nu_c = 0.2
C = 35
A_t = 0.7 # 填入A_t的值
B_t = 10000 # 填入B_t的值
w_surface = 0.3
epsilon_t = 2.2/31500
epsilon_theta = epsilon_t + (w_surface / (2 * np.pi * (R + C)))
term1 = 1 - (epsilon_t * (1 - A_t) / epsilon_theta)
term2 = A_t / np.exp(B_t * (epsilon_theta - epsilon_t))
D_N = term1 - term2
# Define the neural network
class PINN(tf.keras.Model):
    def __init__(self):
        super(PINN, self).__init__()
        self.dense1 = tf.keras.layers.Dense(10, activation=tf.nn.tanh, input_shape=(1,))
        self.dense2 = tf.keras.layers.Dense(50, activation=tf.nn.tanh)
        self.dense3 = tf.keras.layers.Dense(100, activation=tf.nn.tanh)
        self.dense4 = tf.keras.layers.Dense(20, activation=tf.nn.tanh)
        self.dense5 = tf.keras.layers.Dense(20, activation=tf.nn.tanh)
        self.dense6 = tfpl.DenseReparameterization(1)


    def call(self, inputs):
        x = tf.convert_to_tensor(inputs)
        x = self.dense1(x)
        x = self.dense2(x)
        x = self.dense3(x)
        x = self.dense4(x)
        x = self.dense5(x)
        x = self.dense6(x)
        return x

# Define the loss function
def loss(model, x, y, boundary_x, boundary_y):
    with tf.GradientTape(persistent=True) as tape:
        with tf.GradientTape() as tape2:
            tape2.watch(x)  # Ensure that x is watched by the tape for gradient calculation
            y_pred = model(x)
 #           negative_penalty = tf.reduce_sum(tf.maximum(0., -y_pred))
            f_pred = tape2.gradient(y_pred, x)# 假设 f_pred 是关于 x 的一阶导数
            # print("x:", x.shape)
 #           print("f_pred", f_pred)
        tape.watch(x)
        f_pred_again = tape.gradient(f_pred, x)
    # Define the differential equation
    # f = f_pred - (0.1213751909167007 * y_pred + 3.6503124665085633) / x
    f = f_pred_again + f_pred/x - ((1-D_N) * y_pred / x ** 2)
    mse = tf.reduce_mean(tf.square(f-boundary_y))  # MSE loss for the differential equation
    print("mse:", mse)

    with tf.GradientTape() as tape4:
        tape4.watch(x)
        h1 = 31500 * (f_pred + 0.2 * (1 - D_N) ** 0.5 * y_pred / x) / 0.96
        h2 = 31500 * (0.2 * (1 - D_N) ** 0.5 * f_pred + (1 - D_N) * y_pred / x) / 0.96
        h1_pred = tape4.gradient(h1, x)
    f2 = h1_pred + (h1 - h2)/x
    z1 = model(boundary_x)
    f1 = z1/boundary_x - 2.2/31500 - 0.3/(2 * np.pi * 45)
    mse2 = tf.reduce_mean(tf.square(f2 - boundary_y))
    mse1 = tf.reduce_mean(tf.square(f1 - boundary_y))
    print("mse1:", mse1)
    with tf.GradientTape() as tape3:
       tape3.watch(boundary_x)
       z1 = model(boundary_x)
#       print(z1)
       z2 = tape3.gradient(z1, boundary_x)
 #      print(z2)
       z3 =  (z2 + 0.2*(1-D_N)**0.5*z1/boundary_x)
    boundary_loss = tf.reduce_mean(tf.square(z3 - boundary_y))  # Boundary condition loss
    print("boundary:", boundary_loss)
    return mse + boundary_loss + mse1

# Define the boundary condition
boundary_x = tf.constant([[45]], dtype=tf.float32)
boundary_y = tf.constant([[0]], dtype=tf.float32)

# Generate training data
x_train = np.linspace(10, 45, 200).reshape(-1, 1).astype(np.float32)
x_train = tf.constant(x_train)
y_train = np.zeros_like(x_train)

# Initialize PINN
model = PINN()

# Define optimization step
optimizer = tf.keras.optimizers.Adam()

# Training loop
epochs = 40000
for epoch in range(epochs):
    with tf.GradientTape() as tape:
        loss_value = loss(model, x_train, y_train, boundary_x, boundary_y)
    gradients = tape.gradient(loss_value, model.trainable_variables)
    optimizer.apply_gradients(zip(gradients, model.trainable_variables))
    if epoch % 100 == 0:
        print("Epoch {}: Loss {}".format(epoch, loss_value.numpy()))

#
# # 保存模型权重
model.save_weights('model.weights.h5')

# 至少调用一次模型，以便 TensorFlow 创建模型的变量
_ = model(x_train[:1])
# 加载模型权重
model.load_weights('model.weights.h5')

# Evaluate the trained model
x_test = np.linspace(10, 45, 100).reshape(-1, 1).astype(np.float32)

# 创建Excel工作簿和工作表
wb = Workbook()
ws = wb.active
ws.title = 'sheet1'

# 写入数据到工作表
for i in range(len(x_test)):
    # 提取x_test中第i行的第0个元素（即单个值）
    value = x_test[i, 0]
    ws.cell(row=i + 1, column=1, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）

# 保存Excel文件
wb.save('depth_BNN1.xlsx')

x_test = tf.constant(x_test)
y_test = np.zeros_like(x_test)

for i in range(len(y_test)):
    # 提取x_test中第i行的第0个元素（即单个值）
    value = y_test[i, 0]
    ws.cell(row=i + 1, column=2, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）

# 保存Excel文件
wb.save('depth_BNN1.xlsx')

import matplotlib.pyplot as plt

# 运行模型多次以获取每个点的预测样本
num_samples = 1000
predictions = [model(x_test).numpy() for _ in range(num_samples)]

# 计算每个点的平均值和标准差
mean_predictions = np.mean(predictions, axis=0)
std_predictions = np.std(predictions, axis=0)
print(x_test.shape)
print(mean_predictions.shape)
print(std_predictions.shape)

x_test = x_test.numpy().flatten()
mean_predictions = mean_predictions.flatten()
std_predictions = std_predictions.flatten()

for i in range(len(mean_predictions)):
    # 提取x_test中第i行的第0个元素（即单个值）
    value = mean_predictions[i]
    ws.cell(row=i + 1, column=3, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）

# 保存Excel文件
wb.save('depth_BNN1.xlsx')

for i in range(len(std_predictions)):
    # 提取x_test中第i行的第0个元素（即单个值）
    value = std_predictions[i]
    ws.cell(row=i + 1, column=4, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）

# 保存Excel文件
wb.save('depth_BNN1.xlsx')

# 绘制结果
plt.plot(x_test, mean_predictions, label='Predicted Mean')
plt.fill_between(x_test, mean_predictions - 2 * std_predictions, mean_predictions + 2 * std_predictions, alpha=0.3,
                 label='Uncertainty')
plt.scatter(boundary_x, boundary_y, color='red', label='Boundary Condition')
plt.xlabel('x')
plt.ylabel('y')
plt.legend()
plt.show()

