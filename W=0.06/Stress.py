import numpy as np
import tensorflow as tf
from scipy.optimize import minimize
import xlwt
from openpyxl import Workbook  
# Define the neural network
class PINN(tf.keras.Model):
    def __init__(self):
        super(PINN, self).__init__()
        self.dense1 = tf.keras.layers.Dense(50, activation=tf.nn.tanh, input_shape=(1,))
        self.dense2 = tf.keras.layers.Dense(50, activation=tf.nn.tanh)
        self.dense3 = tf.keras.layers.Dense(1)

    def call(self, inputs):
        x = tf.convert_to_tensor(inputs)
        x = self.dense1(x)
        x = self.dense2(x)
        x = self.dense3(x)
        return x

# Define the loss function
def loss(model, x, y, boundary_x, boundary_y):
    with tf.GradientTape() as tape:
        tape.watch(x)  # Ensure that x is watched by the tape for gradient calculation
        y_pred = model(x)
        f_pred = tape.gradient(y_pred, x)  # Derivative of y_pred w.r.t. x
        # print("x:", x.shape)
        # print("y_pred_dist:", y_pred.shape)
    # Define the differential equation
    f = f_pred - (0.2976275431312545 * y_pred +  3.9535950547537033) / x
    mse = tf.reduce_mean(tf.square(f))  # MSE loss for the differential equation
    boundary_loss = tf.reduce_mean(tf.square(model(boundary_x) - boundary_y))  # Boundary condition loss
    return mse + boundary_loss

# Define the boundary condition
boundary_x = tf.constant([[45]], dtype=tf.float32)
boundary_y = tf.constant([[0]], dtype=tf.float32)

# Generate training data
x_train = np.linspace(10, 60, 200).reshape(-1, 1).astype(np.float32)
x_train = tf.constant(x_train)
y_train = np.zeros_like(x_train)

# Initialize PINN
model = PINN()

# Define optimization step
optimizer = tf.keras.optimizers.Adam()

# Training loop
epochs = 2000
for epoch in range(epochs):
    with tf.GradientTape() as tape:
        loss_value = loss(model, x_train, y_train, boundary_x, boundary_y)
    gradients = tape.gradient(loss_value, model.trainable_variables)
    optimizer.apply_gradients(zip(gradients, model.trainable_variables))
    if epoch % 100 == 0:
        print("Epoch {}: Loss {}".format(epoch, loss_value.numpy()))

# Evaluate the trained model
x_test = np.linspace(10, 45, 200).reshape(-1, 1).astype(np.float32)
#print(x_test)

# 创建Excel工作簿和工作表  
wb = Workbook()  
ws = wb.active  
ws.title = 'sheet1'  
  
# 写入数据到工作表  
for i in range(len(x_test)):  
    # 提取x_test中第i行的第0个元素（即单个值）  
    value = x_test[i, 0]  
    ws.cell(row=i+1, column=1, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）  
  
# 保存Excel文件  
wb.save('PINN.xlsx')

x_test= tf.constant(x_test)
y_pred = model(x_test).numpy()
#print(y_pred)

for i in range(len(y_pred)):  
    # 提取x_test中第i行的第0个元素（即单个值）  
    value = y_pred[i, 0]  
    ws.cell(row=i+1, column=2, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）  
  
# 保存Excel文件  
wb.save('PINN.xlsx')

x_eval = tf.constant([[10.0]])
y_eval = model(x_eval)
x_30 = tf.constant([[30.0]])
y_30 = model(x_30)
x_20 = tf.constant([[20.0]])
y_20 = model(x_20)
x_40 = tf.constant([[40.0]])
y_40 = model(x_40)
print(f'Predicted value at x=10: {y_eval.numpy()}')
print(f'Predicted value at x=20: {y_20.numpy()}')
print(f'Predicted value at x=30: {y_30.numpy()}')
print(f'Predicted value at x=40: {y_40.numpy()}')

# Plot the results
import matplotlib.pyplot as plt
plt.plot(x_test, y_pred, label='Predicted')
plt.scatter(boundary_x, boundary_y, color='red', label='Boundary Condition')
plt.xlabel('x')
plt.ylabel('y')
plt.legend()
plt.show()

