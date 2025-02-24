import numpy as np
import tensorflow as tf
from scipy.optimize import minimize
import xlwt
from openpyxl import Workbook
R = 10
n0 = 2/1.1
E_c = 31500
q = 4.76
delta_c = 0.8
nu_c = 0.2
C = 35
A_t = 0.7 # 填入A_t的值
B_t = 10000 # 填入B_t的值
w_surface = 0.06
epsilon_t = 2.2/31500
epsilon_theta = epsilon_t + (w_surface / (2 * np.pi * (R + C)))
term1 = 1 - (epsilon_t * (1 - A_t) / epsilon_theta)
term2 = A_t / np.exp(B_t * (epsilon_theta - epsilon_t))
D_N = term1 - term2
# Define the neural network
class PINN(tf.keras.Model):
    def __init__(self):
        super(PINN, self).__init__()
        self.dense1 = tf.keras.layers.Dense(10, activation=tf.nn.tanh, input_shape=(1,))
        self.dense2 = tf.keras.layers.Dense(50, activation=tf.nn.tanh)
        self.dense3 = tf.keras.layers.Dense(100, activation=tf.nn.tanh)
        self.dense4 = tf.keras.layers.Dense(20, activation=tf.nn.tanh)
        self.dense5 = tf.keras.layers.Dense(20, activation=tf.nn.tanh)
        self.dense6 = tf.keras.layers.Dense(1)

    def call(self, inputs):
        x = tf.convert_to_tensor(inputs)
        x = self.dense1(x)
        x = self.dense2(x)
        x = self.dense3(x)
        x = self.dense4(x)
        x = self.dense5(x)
        x = self.dense6(x)
        return x

# Define the loss function
def loss(model, x, y, boundary_x, boundary_y):
    with tf.GradientTape(persistent=True) as tape:
        with tf.GradientTape() as tape2:
            tape2.watch(x)  # Ensure that x is watched by the tape for gradient calculation
            y_pred = model(x)
            negative_penalty = tf.reduce_sum(tf.maximum(0., -y_pred))
            f_pred = tape2.gradient(y_pred, x)# 假设 f_pred 是关于 x 的一阶导数
            # print("x:", x.shape)
 #           print("f_pred", f_pred)
        tape.watch(x)
        f_pred_again = tape.gradient(f_pred, x)
    # Define the differential equation
    # f = f_pred - (0.1213751909167007 * y_pred + 3.6503124665085633) / x
    f = f_pred_again + f_pred/x - ((1-D_N) * y_pred / x ** 2)
    mse = tf.reduce_mean(tf.square(f-boundary_y))  # MSE loss for the differential equation
    print("mse:", mse)

    with tf.GradientTape() as tape4:
        tape4.watch(x)
        h1 = 31500 * (f_pred + 0.2 * (1 - D_N) ** 0.5 * y_pred / x) / 0.96
        h2 = 31500 * (0.2 * (1 - D_N) ** 0.5 * f_pred + (1 - D_N) * y_pred / x) / 0.96
        h1_pred = tape4.gradient(h1, x)
    f2 = h1_pred + (h1 - h2)/x
    z1 = model(boundary_x)
    f1 = z1/boundary_x - 2.2/31500 - 0.06/(2 * np.pi * 45)
    mse2 = tf.reduce_mean(tf.square(f2 - boundary_y))
    mse1 = tf.reduce_mean(tf.square(f1 - boundary_y))
    print("mse1:", mse1)
    with tf.GradientTape() as tape3:
       tape3.watch(boundary_x)
       z1 = model(boundary_x)
#       print(z1)
       z2 = tape3.gradient(z1, boundary_x)
 #      print(z2)
       z3 =  (z2 + 0.2*(1-D_N)**0.5*z1/boundary_x)
    boundary_loss = tf.reduce_mean(tf.square(z3 - boundary_y))  # Boundary condition loss
    print("boundary:", boundary_loss)
    return mse + boundary_loss + mse1 + negative_penalty

# Define the boundary condition
boundary_x = tf.constant([[45]], dtype=tf.float32)
boundary_y = tf.constant([[0]], dtype=tf.float32)

# Generate training data
x_train = np.linspace(10, 45, 200).reshape(-1, 1).astype(np.float32)
x_train = tf.constant(x_train)
y_train = np.zeros_like(x_train)

# Initialize PINN
model = PINN()

# Define optimization step
optimizer = tf.keras.optimizers.Adam()

# Training loop
epochs = 20000
for epoch in range(epochs):
    with tf.GradientTape() as tape:
        loss_value = loss(model, x_train, y_train, boundary_x, boundary_y)
    gradients = tape.gradient(loss_value, model.trainable_variables)
    optimizer.apply_gradients(zip(gradients, model.trainable_variables))
    if epoch % 100 == 0:
        print("Epoch {}: Loss {}".format(epoch, loss_value.numpy()))

# Evaluate the trained model
x_test = np.linspace(10, 45, 200).reshape(-1, 1).astype(np.float32)
#print(x_test)

# 创建Excel工作簿和工作表  
wb = Workbook()  
ws = wb.active  
ws.title = 'sheet1'  
  
# 写入数据到工作表  
for i in range(len(x_test)):  
    # 提取x_test中第i行的第0个元素（即单个值）  
    value = x_test[i, 0]  
    ws.cell(row=i+1, column=1, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）  
  
# 保存Excel文件  
wb.save('PINN1.xlsx')

x_test= tf.constant(x_test)
y_pred = model(x_test).numpy()
#print(y_pred)


for i in range(len(y_pred)):  
    # 提取x_test中第i行的第0个元素（即单个值）  
    value = y_pred[i, 0]  
    ws.cell(row=i+1, column=2, value=value)  # 写入数据到第i+1行第2列（因为行和列索引都从1开始）  
  
# 保存Excel文件  
wb.save('PINN1.xlsx')

#x_eval = tf.constant([[10.0]])
#y_eval = model(x_eval)
#x_30 = tf.constant([[30.0]])
#y_30 = model(x_30)
#x_20 = tf.constant([[20.0]])
#y_20 = model(x_20)
x_445 = tf.constant([[44.5]])
y_445 = model(x_445)
#print(f'Predicted value at x=10: {y_eval.numpy()}')
#print(f'Predicted value at x=20: {y_20.numpy()}')
#print(f'Predicted value at x=30: {y_30.numpy()}')
print(f'Predicted value at x=40: {y_445.numpy()}')
with tf.GradientTape() as tape4:
    tape4.watch(x_445)
    z1 = model(x_445)
#   print(z1)
    epsilon_r = tape4.gradient(z1, x_445)
    epsilon = z1 / x_test
    print(f'epsilon_r: {epsilon_r}')
    z = 31500 * (epsilon_r + 0.2*(1-D_N)**0.5*epsilon)/0.96


# Plot the results
import matplotlib.pyplot as plt
plt.plot(x_test, y_pred, label='Predicted')
plt.scatter(boundary_x, boundary_y, color='red', label='Boundary Condition')
plt.xlabel('x')
plt.ylabel('y')
plt.legend()
plt.show()

plt.plot(x_test, z, label='Predicted')
plt.scatter(boundary_x, boundary_y, color='red', label='Boundary Condition')
plt.xlabel('x')
plt.ylabel('y')
plt.legend()
plt.show()


